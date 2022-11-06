import pandas as pd
import os
from openpyxl import load_workbook
from shapely import wkt
import geopandas as gpd

def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

# Use wkt to describe a polygon (in Australia albers (3577)) for clipping our NGIS data

bounding_box = "Polygon ((1017760.76728096301667392 -3541957.70058736298233271, 1033044.96912370761856437 -3380500.8970661829225719, 1362219.94902850710786879 -3210531.94267989695072174, 1461582.75826686783693731 -3229754.48591392068192363, 1547489.18534193048253655 -3368580.82105459086596966, 1518569.3516713276039809 -3414031.0364407654851675, 1418155.4332151913549751 -3421778.12315146625041962, 1359002.49710371834225953 -3370541.71342544769868255, 1317053.06696818908676505 -3416099.50407753139734268, 1073812.87129989336244762 -3543764.89405762869864702, 1017760.76728096301667392 -3541957.70058736298233271))"

p1 = wkt.loads(bounding_box)

infile = r"E:\GA\UDF\compilation\staging\UDAEM_HYSCRIPT_DataSplitInTabs_290421.xlsx"

sheets = get_sheetnames_xlsx(infile)

df_DPIE = pd.read_excel(infile, engine = "openpyxl", sheet_name = 'Site')

gdf = gpd.GeoDataFrame(df_DPIE, geometry=gpd.points_from_xy(df_DPIE.Longitude, df_DPIE.Latitude), crs = "epsg:4283").to_crs("epsg:3577")

gdf_DPIE = gdf[gdf.geometry.within(p1)]

infile2 = r"E:\GA\UDF\compilation\staging\UDF_bore_loading.xlsx"

df_NGIS = pd.read_excel(infile2, engine = "openpyxl", sheet_name = 'UDF_Bore')

bores = df_NGIS['StateBoreID'].values

mask = gdf_DPIE['Site'].isin(bores)

# get bores not in NGIS

missing_bores = gdf_DPIE[~mask]['Site'].values

outfile = r"E:\GA\UDF\compilation\staging\UDAEM_HYSCRIPT_site.csv"

gdf_DPIE[~mask].to_csv(outfile, index = False)

for sheet in sheets[1:]:
    print(sheet)
    df_sheet = pd.read_excel(infile, engine="openpyxl", sheet_name=sheet)
    mask = df_sheet['Site'].isin(missing_bores)
    outfile = r"E:\GA\UDF\compilation\staging\UDAEM_HYSCRIPT_{}.csv".format(sheet)
    if len(df_sheet[mask]) != 0:
        df_sheet[mask].to_csv(outfile)
